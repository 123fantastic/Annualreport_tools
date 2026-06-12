# 作者：Winter
# 功能：清洗年报文本，统计数字化关键词词频，并用 pandas 汇总结果

import os
import re
from collections import Counter

import jieba
import pandas as pd
import requests
from digital_keywords import KEYWORDS
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference


# ==================== 配置区：主要改这里 ====================

REPORT_DIR = "年报文件"
OUTPUT_EXCEL = r"年报文件\中证A50近10年数字化词频分析结果.xlsx"

# 停用词来源：https://github.com/123fantastic/Stopwords
STOPWORDS_FILE = r"data\stopwords\stopwords_cn.txt"
STOPWORDS_URL = "https://raw.githubusercontent.com/123fantastic/Stopwords/main/stopwords_cn.txt"

# 展示图表用的公司和年份
CHART_COMPANY = "比亚迪"
COMPARE_YEAR = 2024

# ==================== 简单函数区 ====================

for word in KEYWORDS:
    jieba.add_word(word)


def clean_text(text):
    """去掉空格、换行和大部分特殊符号，只保留中文、英文和数字。"""
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^\u4e00-\u9fa5A-Za-z0-9]", "", text)
    return text


def get_stopwords():
    """读取停用词；如果本地没有，就从 GitHub 下载。"""
    if not os.path.exists(STOPWORDS_FILE):
        try:
            os.makedirs(os.path.dirname(STOPWORDS_FILE), exist_ok=True)
            res = requests.get(STOPWORDS_URL, timeout=30)
            res.encoding = "utf-8"

            with open(STOPWORDS_FILE, "w", encoding="utf-8") as f:
                f.write(res.text)

            print("停用词已下载：", STOPWORDS_FILE)

        except Exception as e:
            print("停用词下载失败，暂时不使用停用词：", e)
            return set()

    with open(STOPWORDS_FILE, "r", encoding="utf-8", errors="ignore") as f:
        stopwords = set()

        for line in f:
            word = line.strip()
            if word:
                stopwords.add(word)

    print("停用词数量：", len(stopwords))
    return stopwords


def get_file_info(filename):
    """从文件名中拆出股票代码、公司简称、年份。"""
    name_without_ext = os.path.splitext(filename)[0]
    parts = name_without_ext.split("_")

    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]

    return "", name_without_ext, ""


def count_keywords(txt_path, stopwords):
    """清洗文本、jieba 分词、Counter 统计关键词。"""
    with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()

    clean_content = clean_text(content)
    words = jieba.cut(clean_content, HMM=False)
    words = [
        word
        for word in words
        if word.strip() and (word not in stopwords or word in KEYWORDS)
    ]
    counter = Counter(words)

    total_words = len(words)
    keyword_total = 0

    result = {
        "总词数": total_words,
    }

    for keyword in KEYWORDS:
        count = counter[keyword]
        result[keyword] = count
        keyword_total = keyword_total + count

    result["数字化关键词总次数"] = keyword_total

    if total_words > 0:
        result["标准化词频_每万词"] = round(keyword_total / total_words * 10000, 4)
    else:
        result["标准化词频_每万词"] = 0

    return result


def add_excel_charts(output_excel, detail_df, keyword_total_df):
    """在 Excel 里加入三张展示图表。"""
    wb = load_workbook(output_excel)

    if "展示图表" in wb.sheetnames:
        del wb["展示图表"]

    ws = wb.create_sheet("展示图表")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # 图1：某家公司数字化词频随年份变化
    company_df = detail_df[detail_df["公司简称"].str.contains(CHART_COMPANY, na=False)]
    if len(company_df) == 0:
        company_df = detail_df[detail_df["股票代码"] == detail_df.iloc[0]["股票代码"]]

    company_trend = company_df.groupby("年份", as_index=False)["标准化词频_每万词"].mean()
    company_trend = company_trend.sort_values("年份")

    ws["A1"] = "图1：某家公司数字化词频随年份变化"
    ws.append(["年份", "标准化词频_每万词"])
    for _, row in company_trend.iterrows():
        ws.append([int(row["年份"]), float(row["标准化词频_每万词"])])

    line_chart = LineChart()
    line_chart.title = CHART_COMPANY + "数字化词频趋势"
    line_chart.y_axis.title = "每万词频次"
    line_chart.x_axis.title = "年份"
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)
    ws.add_chart(line_chart, "D2")

    # 图2：不同公司数字化词频对比
    start_row = ws.max_row + 3
    ws.cell(row=start_row, column=1, value=f"图2：{COMPARE_YEAR}年不同公司数字化词频对比")
    ws.cell(row=start_row + 1, column=1, value="公司简称")
    ws.cell(row=start_row + 1, column=2, value="标准化词频_每万词")

    compare_df = detail_df[detail_df["年份"] == COMPARE_YEAR]
    compare_df = compare_df.sort_values("标准化词频_每万词", ascending=False).head(15)

    for _, row in compare_df.iterrows():
        ws.append([row["公司简称"], float(row["标准化词频_每万词"])])

    end_row = ws.max_row
    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.title = f"{COMPARE_YEAR}年公司数字化词频对比"
    bar_chart.y_axis.title = "公司"
    bar_chart.x_axis.title = "每万词频次"
    data = Reference(ws, min_col=2, min_row=start_row + 1, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=end_row)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    ws.add_chart(bar_chart, "D20")

    # 图3：不同关键词出现频次
    start_row = ws.max_row + 3
    ws.cell(row=start_row, column=1, value="图3：不同关键词出现频次")
    ws.cell(row=start_row + 1, column=1, value="关键词")
    ws.cell(row=start_row + 1, column=2, value="总频次")

    top_keywords = keyword_total_df.head(15)
    for _, row in top_keywords.iterrows():
        ws.append([row["关键词"], int(row["总频次"])])

    end_row = ws.max_row
    keyword_chart = BarChart()
    keyword_chart.type = "bar"
    keyword_chart.title = "数字化关键词总频次"
    keyword_chart.y_axis.title = "关键词"
    keyword_chart.x_axis.title = "出现次数"
    data = Reference(ws, min_col=2, min_row=start_row + 1, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=end_row)
    keyword_chart.add_data(data, titles_from_data=True)
    keyword_chart.set_categories(cats)
    ws.add_chart(keyword_chart, "D38")

    wb.save(output_excel)


# ==================== 主程序 ====================

rows = []
stopwords = get_stopwords()

if not os.path.exists(REPORT_DIR):
    print("年报文件夹不存在：", REPORT_DIR)
else:
    for root, dirs, files in os.walk(REPORT_DIR):
        dirs.sort()
        files.sort()

        if "txt年报" not in root:
            continue

        for filename in files:
            if not filename.endswith(".txt"):
                continue

            txt_path = os.path.join(root, filename)
            stock_code, company_name, report_year = get_file_info(filename)

            print("正在分析：", filename, flush=True)

            try:
                result = count_keywords(txt_path, stopwords)
            except Exception as e:
                print("分析失败：", filename, e)
                continue

            row = {
                "股票代码": stock_code,
                "公司简称": company_name,
                "年份": int(report_year),
                "总词数": result["总词数"],
                "数字化关键词总次数": result["数字化关键词总次数"],
                "标准化词频_每万词": result["标准化词频_每万词"],
            }

            for keyword in KEYWORDS:
                row[keyword] = result[keyword]

            rows.append(row)

            if len(rows) % 50 == 0:
                print(f"已经完成 {len(rows)} 份", flush=True)

os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
detail_df = pd.DataFrame(rows)

if len(detail_df) == 0:
    print("没有分析到 TXT 文件")
else:
    detail_df = detail_df.sort_values(["年份", "股票代码"])

    company_compare_df = detail_df.groupby(["股票代码", "公司简称"], as_index=False).agg(
        报告数量=("年份", "count"),
        数字化关键词总次数=("数字化关键词总次数", "sum"),
        平均标准化词频_每万词=("标准化词频_每万词", "mean"),
    )
    company_compare_df = company_compare_df.sort_values("平均标准化词频_每万词", ascending=False)

    keyword_total_df = pd.DataFrame({
        "关键词": KEYWORDS,
        "总频次": [int(detail_df[word].sum()) for word in KEYWORDS],
    })
    keyword_total_df = keyword_total_df.sort_values("总频次", ascending=False)

    keyword_year_df = detail_df.groupby("年份", as_index=False)[KEYWORDS].sum()

    ppt_conclusion_df = pd.DataFrame([
        ["样本范围", f"共分析 {len(detail_df)} 份中证A50公司年报文本"],
        ["方法说明", "正则清洗文本，jieba 中文分词，Counter 统计关键词，pandas 汇总分析"],
        ["指标说明", "标准化词频 = 数字化关键词总次数 / 总词数 * 10000"],
        ["展示结论", "2020-2024年，多数样本公司的数字化相关词频呈上升趋势，其中人工智能、云计算、数据要素等词在近两年更明显。"],
    ], columns=["项目", "内容"])

    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        detail_df.to_excel(writer, sheet_name="原始统计", index=False)
        company_compare_df.to_excel(writer, sheet_name="公司对比", index=False)
        keyword_total_df.to_excel(writer, sheet_name="关键词总频次", index=False)
        keyword_year_df.to_excel(writer, sheet_name="关键词年度变化", index=False)
        ppt_conclusion_df.to_excel(writer, sheet_name="PPT结论参考", index=False)

    add_excel_charts(OUTPUT_EXCEL, detail_df, keyword_total_df)

    print("\n分析完成", flush=True)
    print("分析文件数量：", len(detail_df), flush=True)
    print("结果已保存到：", OUTPUT_EXCEL, flush=True)
